#!/usr/bin/env python3
"""
StakTrakr Copper Pre-2013 History Importer (STRK-304)
=====================================================
One-shot importer for the World Bank "Pink Sheet" monthly commodity series.
Fills the copper history gap BEFORE MetalPriceAPI's XCU floor (2013-07-23),
exactly the way the LBMA dataset seeded pre-API gold/silver history.

Series: "Copper" — LME grade A, minimum 99.9935% purity, cathodes — quoted in
USD per metric tonne, monthly average. Converted to USD per troy ounce,
calibrated onto the API quote basis (see CALIBRATION_TO_API), and written into
data/spot-history-{YEAR}.json as SPARSE monthly points (one entry per month,
timestamped on the 1st at noon). Monthly points are deliberately NOT
forward-filled to daily — that would fabricate precision the source lacks.

Provider tag is "LME-WB" so provenance stays distinguishable from "LBMA"
(gold/silver deep history) and "StakTrakr" (MetalPriceAPI era).

Usage:
    python3 import-copper-pinksheet.py --xlsx /path/to/CMO-Historical-Data-Monthly.xlsx
    python3 import-copper-pinksheet.py --xlsx ... --dry-run
    python3 import-copper-pinksheet.py --xlsx ... --start-month 1968-01 --end-month 2013-06

The xlsx is republished periodically at an edition-hashed URL. Resolve the
current edition from https://www.worldbank.org/en/research/commodity-markets
(link text "Monthly prices") rather than hardcoding a URL here.
Edition used for the STRK-304 import: "Updated on August 04, 2026"
(74e8be41ceb20fa0da750cda2f6b9e4e-0050012026).

Requires: openpyxl (not part of the poller runtime deps — install ad hoc).
"""

import argparse
import importlib.util
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Error: openpyxl is required to parse the Pink Sheet xlsx but is not "
        "installed (it is deliberately not a poller runtime dependency). "
        "Install it ad hoc: pip install openpyxl"
    )

# Reuse the year-file merge + rounding contract from the seed updater so the
# two writers can never drift apart on shape, dedup, or precision rules.
# (The filename is hyphenated, so load it via importlib rather than import.)
_SEED_PATH = Path(__file__).resolve().parent / "update-seed-data.py"
_spec = importlib.util.spec_from_file_location("update_seed_data", _SEED_PATH)
_seed = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_seed)
merge_into_year_files = _seed.merge_into_year_files
resolve_data_dir = _seed.resolve_data_dir
round_price = _seed.round_price

# 1 metric tonne = 1,000,000 g / 31.1034768 g per troy oz
TROY_OZ_PER_TONNE = 32150.7466

# Splice calibration to the app's quote basis (STRK-304).
# The WB/LME series runs systematically ~9% above MetalPriceAPI's XCU quote —
# measured across all 157 overlap months (2013-07 → 2026-07): median ratio
# 1.0944, stdev 0.026, era medians 1.092–1.099 for 2013–2024. That is a
# benchmark-basis difference (LME grade-A cathode settlement vs the API's
# composite), not noise, so imported values are scaled onto the API basis.
# Without this the chart shows a permanent ~9% step at the 2013 splice.
# Derivation: 1 / 1.0944. Boundary check: WB 2013-06 $0.2177/ozt × 0.9138 =
# $0.1989 vs the API's first day (2013-07-23) $0.1999 — a 0.5% seam.
CALIBRATION_TO_API = 0.9138

SHEET_NAME = "Monthly Prices"
HEADER_ROW = 5  # 1-based: commodity names
UNIT_ROW = 6  # 1-based: units, must read "($/mt)" for Copper
DATA_START_ROW = 7
COPPER_HEADER = "Copper"
EXPECTED_UNIT = "($/mt)"


def parse_month_token(token):
    """Parse a Pink Sheet month token like '1968M01' into ('1968', '01') or None."""
    if not isinstance(token, str) or "M" not in token:
        return None
    year, _, month = token.partition("M")
    if len(year) == 4 and year.isdigit() and len(month) == 2 and month.isdigit():
        return year, month
    return None


def find_copper_column(ws):
    """Locate the Copper column index (1-based) and assert its unit is $/mt."""
    headers = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    units = next(ws.iter_rows(min_row=UNIT_ROW, max_row=UNIT_ROW, values_only=True))
    for idx, name in enumerate(headers):
        if isinstance(name, str) and name.strip() == COPPER_HEADER:
            unit = (units[idx] or "").strip() if idx < len(units) else ""
            if unit != EXPECTED_UNIT:
                raise RuntimeError(
                    f"Copper column unit is {unit!r}, expected {EXPECTED_UNIT!r} — "
                    "the sheet layout changed; refusing to convert blindly."
                )
            return idx + 1
    raise RuntimeError(f"No column named {COPPER_HEADER!r} in sheet {SHEET_NAME!r}.")


def extract_entries(xlsx_path, start_month, end_month):
    """Read monthly copper prices and return seed-format entries in $/ozt."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Sheet {SHEET_NAME!r} not found in {xlsx_path}.")
        ws = wb[SHEET_NAME]
        copper_col = find_copper_column(ws)

        entries = []
        skipped = 0
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            parsed = parse_month_token(row[0])
            if not parsed:
                continue
            year, month = parsed
            ym = f"{year}-{month}"
            if ym < start_month or ym > end_month:
                continue
            raw = row[copper_col - 1]
            try:
                per_tonne = float(raw)
            except (TypeError, ValueError):
                skipped += 1  # "…" placeholder or blank — no observation that month
                continue
            entries.append(
                {
                    "spot": round_price(per_tonne / TROY_OZ_PER_TONNE * CALIBRATION_TO_API),
                    "metal": "Copper",
                    "source": "seed",
                    "provider": "LME-WB",
                    "timestamp": f"{ym}-01 12:00:00",
                }
            )
        return entries, skipped
    finally:
        wb.close()


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import pre-2013 monthly copper history from the World Bank Pink Sheet."
    )
    parser.add_argument("--xlsx", required=True, help="Path to CMO-Historical-Data-Monthly.xlsx")
    parser.add_argument(
        "--start-month",
        default="1968-01",
        help="First month to import, YYYY-MM (default 1968-01 — the app's history baseline).",
    )
    parser.add_argument(
        "--end-month",
        default="2013-06",
        help=(
            "Last month to import, YYYY-MM (default 2013-06 — the API's daily XCU "
            "series starts 2013-07-23 and owns everything after)."
        ),
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing.")
    return parser.parse_args()


def main():
    args = parse_args()
    data_dir = resolve_data_dir()

    print("Pink Sheet Copper Importer (STRK-304)")
    print("=====================================")
    print(f"Source:  {args.xlsx}")
    print(f"Window:  {args.start_month} → {args.end_month} (monthly, sparse)")

    entries, skipped = extract_entries(args.xlsx, args.start_month, args.end_month)
    if not entries:
        print("No copper observations found in the window — nothing to do.")
        sys.exit(1)

    print(f"Parsed:  {len(entries)} monthly observations ({skipped} months without data)")
    print(f"Range:   {entries[0]['timestamp'][:10]} ${entries[0]['spot']} → "
          f"{entries[-1]['timestamp'][:10]} ${entries[-1]['spot']}")

    results = merge_into_year_files(data_dir, entries, dry_run=args.dry_run)

    print("Updated files:" if not args.dry_run else "Would update files:")
    for year, count in sorted(results.items()):
        print(f"  spot-history-{year}.json: +{count} entries")
    total = sum(results.values())
    print(f"\nDone. {total} entries {'added' if not args.dry_run else 'would be added'}.")


if __name__ == "__main__":
    main()
